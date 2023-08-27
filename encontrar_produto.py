# Importe a biblioteca os para manipular o sistema de arquivos
import os
# Importe a biblioteca BeautifulSoup para análise HTML
from bs4 import BeautifulSoup
# Importe a biblioteca requests para fazer requisições HTTP
import requests
# Importe a biblioteca openpyxl para trabalhar com arquivos Excel
import openpyxl
# Importe o módulo datetime para lidar com datas
from datetime import datetime

# Nome do que quer procurar no site
Nome_ItemProcurado = "RTX"
# Nome do arquivo Excel
filename = 'ItensKabum.xlsx'

# Obtenha o caminho da pasta "Documentos" no Windows
documents_folder = os.path.join(os.environ["USERPROFILE"], "Documents")

# Verifique se o arquivo já existe na pasta "Documentos"
if not os.path.exists(os.path.join(documents_folder, filename)):
    # Se o arquivo não existe, crie um novo arquivo Excel
    workbook = openpyxl.Workbook()
else:
    # Se o arquivo já existe, carregue-o
    workbook = openpyxl.load_workbook(os.path.join(documents_folder, filename))

# Obtenha a data de hoje no formato "dd.mm.yyyy"
data_hoje = datetime.now().strftime("%d.%m.%Y")

# Verifique se a planilha já existe no arquivo
if data_hoje not in workbook.sheetnames:
    # Crie uma nova planilha com o nome da data de hoje
    new_sheet = workbook.create_sheet(title=data_hoje)
else:
    # Se a planilha já existe, obtenha a referência à planilha existente
    new_sheet = workbook[data_hoje]

# Contador para o número da página
contagem_page = 0

# Contador geral para as linhas da planilha
contagem = 0

# Loop infinito para percorrer as páginas até não haver mais elementos
while True:
    # Incrementa o número da página
    contagem_page += 1
    
    # Exibe o número da página (apenas para fins de acompanhamento)
    print(contagem_page)
    
    # Constrói a URL da página a ser raspada
    url = "https://www.kabum.com.br/busca/"+ Nome_ItemProcurado +"?page_number="
    url += str(contagem_page) + "&page_size=100"
    
    # Baixa o conteúdo da página usando a biblioteca requests
    response = requests.get(url)
    html_content = response.content
    
    # Cria um objeto BeautifulSoup para analisar o conteúdo HTML da página
    soup = BeautifulSoup(html_content, "html.parser")

    # Encontra todos os elementos com o atributo 'data-testid' igual a 'tooltip'
    all_elements = soup.find_all(attrs={"data-testid": "tooltip"})
    
    # Encontra todos os elementos com a classe 'sc-3b515ca1-2 chPrxA priceCard'
    all_values = soup.find_all(attrs={"class": "sc-3b515ca1-2 chPrxA priceCard"})
    
    # Verifica se a lista de elementos não está vazia
    if len(all_elements) > 0:
        # Loop para percorrer elementos e valores simultaneamente usando zip()
        for element, value in zip(all_elements, all_values):
            # Incrementa o contador de linhas
            contagem += 1
            # Adiciona o texto do elemento na coluna A da planilha
            new_sheet['A' + str(contagem)] = element.text
            # Adiciona o texto do valor na coluna B da planilha
            new_sheet['B' + str(contagem)] = value.text
    else:
        # Se não houver mais elementos, sai do loop while
        break
# Salva o arquivo Excel com os dados coletados na pasta "Documentos"
workbook.save(os.path.join(documents_folder, filename))
